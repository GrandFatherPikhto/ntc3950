import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='./data/NTC3950.xlsx')
# print(wb['2'].cell(row=2, column=2).value)
ws = wb['2']
table = ws['A2':'B128']
temperature = []
resistance = []
# for x in range(0, len(table)):
#     temperature.append(table[x][0].value)
#     resistance.append(table[x][1].value)

for x in range(2, 129):
    temperature.append(ws.cell(row=x, column=1).value)
    resistance.append(ws.cell(row=x, column=2).value)

print(temperature)
print(resistance)
