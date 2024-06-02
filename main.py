# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from openpyxl import load_workbook

def read_xslx(filename, sheet, square):
    wb = load_workbook(filename=filename)
    ws = wb[sheet]
    temperature = []
    resistance = []
    for x in range(2, 129):
        if ws.cell(row=x, column=1).value is None:
            break
        if ws.cell(row=x, column=2).value is None:
            break
        temperature.append(ws.cell(row=x, column=1).value+273.15)
        resistance.append(ws.cell(row=x, column=2).value*1000)
    print(temperature)
    print(resistance)
    return [temperature, resistance]


# Стейнхарта-Харта без деления 1 на температуру
def steinhart(x, a, b, c):
    """Fit function y=f(x,p) with parameters p=(a,b,c). """
    return a + b * np.log(x) + c * np.log(x) * np.log(x) * np.log(x)


# 1.0/T
def steinhart_orig(x, a, b, c):
    return 1.0 / steinhart(x, a, b, c)

# Создать каталог для графиков, если его нет
def make_output_dir (path):
    if not os.path.exists(path):
        os.makedirs(path)


# Строим оригинальный график и интерполированный
def plot_graphs(resistance, temperature, parameters, plotname):
    fitted = steinhart_orig(resistance, *parameters)

    # # Построение графика
    fig, ax = plt.subplots(constrained_layout=True)

    for i, t in enumerate(temperature):
        temperature[i] = t - 273.15
    ax.plot(temperature, resistance, label='Исходные данные NTC 3950')

    for i, t in enumerate(fitted):
        fitted[i] = t - 273.15
    ax.plot(fitted, resistance, label='Интерполяция термистора NTC 3950')

    ax.set(xlabel='Температура °С', ylabel='Сопротивление Ом', title='NTC 3950')
    ax.grid()
    fig.set_size_inches(18.5, 10.5)
    fig.legend()
    fig.savefig(plotname, dpi=600)

    plt.legend()
    plt.show()

    plt.clf()
    plt.cla()


#  Печать коэффициентов A,B,C уравнения Стейнхарта-Харта
def print_params(parameters):
    print('\n== Steinhart-Hart A,B,C parameters ==============================')
    print('#define STEINHART_A\t{A}'.format(A=parameters[0]))
    print('#define STEINHART_B\t{B}'.format(B=parameters[1]))
    print('#define STEINHART_C\t{C}'.format(C=parameters[2]))
    print('\n===End parameters=============================')


# Интерполяция графика
def calc_params(resistance, temperature):
    rtemp = []
    for t in temperature:
        rtemp.append(1.0 / t)
    return curve_fit(steinhart, resistance, rtemp)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    make_output_dir('./plots/')
    temperature, resistance = read_xslx(filename='./data/NTC3950.xlsx', sheet='Sheet3', square='A2:B143')
    parameters, covariance = calc_params(resistance=resistance, temperature=temperature)
    print_params(parameters=parameters)
    plot_graphs(resistance=resistance, temperature=temperature, parameters=parameters, plotname='./plots/ntc3950.png')

